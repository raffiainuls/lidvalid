"""Excel export — kept for transition/export only (PRD FR-14, US-17), not as
the platform's primary output anymore. Ported from
`validation-data/scripts/run_batch.py` (`_write_excel`, `_apply_style`,
`_sheet_name`, `_strip_tz`) merged with the rollup idea from
`validation_database/make_summary_excel.py`.
"""
from __future__ import annotations

import datetime as _dt

import pandas as pd

from .runner.tiered import TableRunResult


def _sheet_name(target_table: str, suffix: str) -> str:
    max_tbl = 31 - len(suffix)
    return f"{target_table[:max_tbl]}{suffix}"


def _strip_tz(df: pd.DataFrame) -> pd.DataFrame:
    """Strip timezone info from datetime-like values so openpyxl can write them."""
    df = df.copy()
    for col in df.columns:
        if hasattr(df[col], "dt") and getattr(df[col].dt, "tz", None) is not None:
            try:
                df[col] = df[col].dt.tz_convert(None)
                continue
            except Exception:
                pass
        if df[col].dtype == object:
            try:
                df[col] = df[col].apply(
                    lambda v: v.replace(tzinfo=None)
                    if isinstance(v, _dt.datetime) and v.tzinfo is not None else v
                )
            except Exception:
                pass
    return df


def build_summary_dataframe(results: list[TableRunResult]) -> pd.DataFrame:
    rows = []
    for r in results:
        agg = r.aggregate_summary or {}
        rl = r.rowlevel
        rows.append({
            "target_table": r.table.target_table,
            "status": r.status,
            "tier_reached": r.tier_reached,
            "mode": r.mode,
            "source_rows": agg.get("source_rows"),
            "target_rows": agg.get("target_rows"),
            "row_diff": agg.get("row_diff"),
            "row_match": agg.get("row_match"),
            "extra_target_columns": ", ".join(agg.get("extra_target_columns") or []) if agg else "",
            "col_completeness_mismatch": agg.get("col_completeness_mismatch"),
            "col_uniqueness_mismatch": agg.get("col_uniqueness_mismatch"),
            "stat_mismatch": agg.get("stat_mismatch"),
            "stat_mismatch_detail": " | ".join(agg.get("stat_mismatch_detail") or []),
            "monthly_mismatch": agg.get("monthly_mismatch"),
            "yearly_mismatch": agg.get("yearly_mismatch"),
            "missing_in_source": rl.missing_in_source_count if rl else None,
            "missing_in_target": rl.missing_in_target_count if rl else None,
            "differing_values": rl.differing_values_count if rl else None,
            "attempts": r.attempts,
            "duration_s": r.duration_s,
            "error": r.error or "",
        })
    return pd.DataFrame(rows)


def write_excel(results: list[TableRunResult], output_path: str) -> None:
    summary_df = build_summary_dataframe(results)

    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        _strip_tz(summary_df).to_excel(writer, sheet_name="Summary", index=False)

        for r in results:
            if r.aggregate is None:
                continue
            tbl = r.table.target_table
            sheets = [
                ("_columns", r.aggregate.column_details),
                ("_src_types", r.aggregate.src_type_details),
                ("_tgt_types", r.aggregate.tgt_type_details),
                ("_monthly", r.aggregate.monthly_breakdown),
                ("_yearly", r.aggregate.yearly_breakdown),
            ]
            for suffix, df in sheets:
                if df is None or df.empty:
                    continue
                sheet = _sheet_name(tbl, suffix)
                _strip_tz(df).reset_index(drop=True).to_excel(writer, sheet_name=sheet, index=False)

            if r.rowlevel and r.rowlevel.differing_values:
                diff_df = pd.DataFrame(r.rowlevel.differing_values)
                _strip_tz(diff_df).to_excel(writer, sheet_name=_sheet_name(tbl, "_diffs"), index=False)

    _apply_style(output_path, summary_df)


def _apply_style(output_path: str, summary_df: pd.DataFrame) -> None:
    """Color Summary rows: green=PASS, yellow=FAIL, red=ERROR — same palette
    as the legacy tool's Excel output."""
    try:
        from openpyxl import load_workbook
        from openpyxl.styles import PatternFill

        fills = {
            "PASS": PatternFill("solid", fgColor="C6EFCE"),
            "FAIL": PatternFill("solid", fgColor="FFEB9C"),
            "ERROR": PatternFill("solid", fgColor="FFC7CE"),
        }
        wb = load_workbook(output_path)
        ws = wb["Summary"]
        status_col = list(summary_df.columns).index("status") + 1  # noqa: F841 (kept for clarity)

        for row_idx, status in enumerate(summary_df["status"], start=2):
            fill = fills.get(status)
            if fill:
                for cell in ws[row_idx]:
                    cell.fill = fill

        for col in ws.columns:
            max_len = max((len(str(c.value or "")) for c in col), default=10)
            ws.column_dimensions[col[0].column_letter].width = min(max_len + 2, 50)

        wb.save(output_path)
        wb.close()
    except Exception as e:  # noqa: BLE001 - styling is cosmetic, never fail the export over it
        print(f"Warning: could not apply Excel styling: {e}")
